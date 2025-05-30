from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
wb = Workbook()
ws = wb.active
heading = ["Roll NO","Name","Math","Science","English"]
ws.append(heading)
student_data = [[101,"Aahil",78,67,84],
                [102,"Sahil",67,77,85],
                [103,"Mohit",89,55,65],
                [104,"Rohit",96,94,88]]

ws.cell(row=1, column=6, value="Total")
ws.cell(row=1, column=7, value="Percentage")
ws.cell(row=1, column=8, value="Grade")
for data in student_data:
    roll_no = data[0]
    name = data[1]
    marks = data[2:]
    total = sum(marks)
    percentage = total/len(marks)
    ws.append([roll_no]+[name]+marks+[total]+[round(percentage,2)])

for row in ws.iter_rows(min_row=2,max_row=5,min_col=7,max_col=7):
    for cell in row:
        percentage = cell.value
        if percentage>=90:
            grade = "A"
        elif percentage>=75:
            grade = "B"
        elif percentage>=60:
            grade ="C"
        else:
            grade = "Fail"


        ws.cell(row=cell.row, column=8, value=grade)
#adjust column width
for coll in ws.columns:
    coll_num =coll[0].column
    coll_letter = get_column_letter(coll_num)
    max_length = 0
    for cell in coll:
        try:
            if cell.value:
                max_length = max(max_length,len(str(cell.value)))
                adjust_length = max_length+2
                ws.column_dimensions[coll_letter].width = adjust_length

        except:
            pass        

#font of first row heading
for row in ws[1]:
    row.font = Font(bold=True)

wb.save("Students.xlsx")